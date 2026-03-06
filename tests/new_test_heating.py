"""
制热模式API测试驱动器
通过HTTP接口模拟用户操作，让前后端正常运行

时间压缩规则：
- Excel中1分钟 = 实际10秒
- 时间压缩比 TIME_SCALE = 6
"""

import os
import sys
import time
import requests
from datetime import datetime
from decimal import Decimal
from openpyxl import load_workbook

# 配置
API_BASE_URL = "http://localhost:8000/api"  # 根据实际后端地址修改
TIME_SCALE = 6
TEST_INTERVAL = 10  # 每行测试数据间隔10秒
DEFAULT_HEATING_TEMP = 25

# 房间初始温度
INITIAL_TEMPS = {
    "301": 10.0,
    "302": 15.0,
    "303": 18.0,
    "304": 12.0,
    "305": 14.0,
}

# 风速映射
FAN_SPEED_MAP = {
    "高": "high",
    "中": "medium",
    "低": "low",
}

# Excel文件路径
TEST_DATA_FILE = os.path.join(os.path.dirname(__file__), "data", "test_hot.xlsx")

# ============================================================
# Excel解析（复用原逻辑）
# ============================================================

def parse_test_data(filepath):
    """解析测试数据Excel"""
    wb = load_workbook(filepath)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[2:]  # 跳过标题
    
    test_actions = []
    
    for row in data_rows:
        if row[0] is None or row[0] == '费用小计':
            continue
            
        time_min = row[0]
        if not isinstance(time_min, (int, float)):
            continue
        
        time_min = int(time_min)
        actions = {}
        
        for room_idx in range(5):
            room_id = f"30{room_idx + 1}"
            cell_value = row[room_idx + 1]
            
            if cell_value is not None:
                action = parse_action(cell_value)
                if action:
                    actions[room_id] = action
        
        if actions or time_min == 0:
            test_actions.append((time_min, actions))
    
    return test_actions


def parse_action(cell_value):
    """解析单元格中的操作指令"""
    if cell_value is None:
        return None
    
    cell_str = str(cell_value).strip()
    
    if cell_str == "开机":
        return {"type": "power_on"}
    elif cell_str == "关机":
        return {"type": "power_off"}
    elif cell_str in FAN_SPEED_MAP:
        return {"type": "change_speed", "fan_speed": FAN_SPEED_MAP[cell_str]}
    elif cell_str.replace(".", "").isdigit():
        return {"type": "change_temp", "target_temp": float(cell_str)}
    elif "，" in cell_str or "," in cell_str:
        parts = cell_str.replace("，", ",").split(",")
        result = {"type": "change_both"}
        for part in parts:
            part = part.strip()
            if part in FAN_SPEED_MAP:
                result["fan_speed"] = FAN_SPEED_MAP[part]
            elif part.replace(".", "").isdigit():
                result["target_temp"] = float(part)
        return result
    elif cell_str == "中央空调启动":
        return {"type": "system_start"}
    elif "检查程序" in cell_str or "设置" in cell_str:
        return None
    
    return None


# ============================================================
# API客户端
# ============================================================

class APIClient:
    def __init__(self, base_url):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
    
    def init_room(self, room_id, temp, mode="heating"):
        """初始化房间温度（仅DEBUG模式）"""
        url = f"{self.base_url}/admin/room/{room_id}/init/"
        response = self.session.post(url, json={"temp": temp, "mode": mode})
        return response.json()
    
    def clear_room(self, room_id):
        """清空房间状态"""
        url = f"{self.base_url}/admin/room/{room_id}/clear/"
        response = self.session.post(url)
        return response.json()

    def checkin(self, room_id, customer_info=None):
        """办理入住"""
        url = f"{self.base_url}/checkin/"
        if not customer_info:
            customer_info = {
                "name": f"测试顾客{room_id}",
                "phone": "13800138000",
                "id_card": f"1234567890{room_id}",
                "room_id": room_id,
            }
        response = self.session.post(url, json=customer_info)
        return response.json()
    
    def control_ac(self, room_id, action, **kwargs):
        """控制空调"""
        url = f"{self.base_url}/ac/control/"
        payload = {"room_id": room_id, "action": action}
        payload.update(kwargs)
        response = self.session.post(url, json=payload)
        return response.json()
    
    def get_ac_state(self, room_id):
        """获取空调状态"""
        url = f"{self.base_url}/ac/state/{room_id}/"
        response = self.session.get(url)
        return response.json()
    
    def get_all_ac_states(self):
        """获取所有房间状态（监控）"""
        url = f"{self.base_url}/ac/monitor/"
        response = self.session.get(url)
        return response.json()
    
    def get_ac_details(self, room_id):
        """获取空调详单"""
        url = f"{self.base_url}/ac/details/{room_id}/"
        response = self.session.get(url)
        return response.json()


# ============================================================
# 测试执行器
# ============================================================

class HeatingAPITest:
    def __init__(self, api_client, test_data_file):
        self.client = api_client
        self.test_data_file = test_data_file
        self.room_ids = ["301", "302", "303", "304", "305"]
        self.room_states = {}
        self.test_start_time = None
        
        # 初始化房间状态缓存
        for room_id in self.room_ids:
            self.room_states[room_id] = {
                "target_temp": DEFAULT_HEATING_TEMP,
                "fan_speed": "medium",
                "is_on": False,
            }
    
    def setup(self):
        """测试环境初始化"""
        print("=" * 60)
        print("制热模式API测试 - 环境初始化")
        print("=" * 60)
        
        # 0. 清空房间状态
        print("\n0. 清空房间状态...")
        for room_id in self.room_ids:
            result = self.client.clear_room(room_id)
            if result.get("code") == 200:
                print(f"  ✅ 房间 {room_id} 状态已清空")
            else:
                print(f"  ⚠️  房间 {room_id} 清空失败: {result.get('message')}")

        # 1. 确保所有房间已入住
        print("\n1. 办理入住...")
        for room_id in self.room_ids:
            result = self.client.checkin(room_id)
            if result.get("code") == 200:
                print(f"  ✅ 房间 {room_id} 入住成功")
            else:
                print(f"  ⚠️  房间 {room_id} 入住失败: {result.get('message')}")
        
        # 2. 初始化房间温度
        print("\n2. 初始化房间温度...")
        for room_id, temp in INITIAL_TEMPS.items():
            result = self.client.init_room(room_id, temp, mode="heating")
            if result.get("code") == 200:
                print(f"  ✅ 房间 {room_id} 初始温度设为 {temp}°C")
            else:
                print(f"  ⚠️  房间 {room_id} 初始化失败: {result.get('message')}")
        
        print("\n环境初始化完成！\n")
    
    def execute_action(self, room_id, action):
        """执行单个操作"""
        action_type = action.get("type")
        
        if action_type == "power_on":
            target_temp = self.room_states[room_id]["target_temp"]
            fan_speed = "medium"  # 开机默认风速
            
            result = self.client.control_ac(
                room_id, "power_on",
                target_temp=target_temp,
                fan_speed=fan_speed,
                mode="heating"
            )
            self.room_states[room_id]["is_on"] = True
            self.room_states[room_id]["fan_speed"] = fan_speed
            print(f"    🔛 开机 (目标{target_temp}°C, {fan_speed})")
            
        elif action_type == "power_off":
            result = self.client.control_ac(room_id, "power_off")
            self.room_states[room_id]["is_on"] = False
            print(f"    ⏹️  关机")
            
        elif action_type == "change_temp":
            target_temp = action.get("target_temp")
            self.room_states[room_id]["target_temp"] = target_temp
            
            # 调温请求直接发送（防抖逻辑）
            result = self.client.control_ac(
                room_id, "change_temp",
                target_temp=target_temp,
                mode="heating"
            )
            print(f"    🌡️  调温 -> {target_temp}°C")
            
        elif action_type == "change_speed":
            fan_speed = action.get("fan_speed")
            self.room_states[room_id]["fan_speed"] = fan_speed
            
            result = self.client.control_ac(
                room_id, "change_speed",
                fan_speed=fan_speed
            )
            print(f"    💨 调风速 -> {fan_speed}")
            
        elif action_type == "change_both":
            target_temp = action.get("target_temp")
            fan_speed = action.get("fan_speed")
            
            if target_temp:
                self.room_states[room_id]["target_temp"] = target_temp
                result = self.client.control_ac(
                    room_id, "change_temp",
                    target_temp=target_temp,
                    mode="heating"
                )
            
            if fan_speed:
                self.room_states[room_id]["fan_speed"] = fan_speed
                result = self.client.control_ac(
                    room_id, "change_speed",
                    fan_speed=fan_speed
                )
            
            print(f"    🔄 调温={target_temp}°C, 调风速={fan_speed}")
        
        # 检查请求是否成功
        if result.get("code") != 200:
            print(f"    ❌ 失败: {result.get('message')}")
        else:
            # 对于power_on/change_speed，更新状态
            if action_type in ["power_on", "change_speed"]:
                self.room_states[room_id].update(result.get("data", {}))
    
    def print_status(self, time_min):
        """打印当前所有房间状态"""
        print(f"\n  📊 [状态] 时间={time_min}分钟")
        print("  " + "-" * 90)
        print(f"  {'房间':<8} {'状态':<12} {'当前温度':<10} {'目标温度':<10} {'风速':<8} {'费用':<8} {'队列':<10}")
        print("  " + "-" * 90)
        
        all_states = self.client.get_all_ac_states().get("data", [])
        state_map = {s["room_id"]: s for s in all_states}
        
        for room_id in self.room_ids:
            state = state_map.get(room_id, {})
            status = state.get("status", "off")
            current = state.get("current_temp", 0)
            target = state.get("target_temp", 0)
            fan_speed = state.get("fan_speed", "-")
            cost = state.get("cost", 0)
            
            # 标记队列位置
            queue_info = ""
            if status == "on":
                queue_info = "[服务]"
            elif status == "waiting":
                remaining = state.get("remaining_wait", 0)
                queue_info = f"[等{remaining:.0f}s]"
            
            print(f"  {room_id:<8} {status:<12} {current:<10.1f} {target:<10.1f} {fan_speed:<8} {cost:<8.2f} {queue_info:<10}")
        
        print("  " + "-" * 90)
    
    def print_final_report(self):
        """打印最终报告"""
        print("\n" + "=" * 60)
        print("测试完成 - 最终报告")
        print("=" * 60)
        
        total_cost = Decimal("0.00")
        total_energy = 0.0
        
        print("\n💰 费用汇总（从详单记录统计）:")
        print("-" * 60)
        
        for room_id in self.room_ids:
            details = self.client.get_ac_details(room_id).get("data", {})
            summary = details.get("summary", {})
            room_cost = Decimal(str(summary.get("total_cost", 0)))
            room_energy = float(summary.get("total_energy", 0))
            
            total_cost += room_cost
            total_energy += room_energy
            
            print(f"  房间 {room_id}: 费用={room_cost:.2f}元, 能耗={room_energy:.2f}度")
        
        print("-" * 60)
        print(f"  总计: 费用={total_cost:.2f}元, 能耗={total_energy:.2f}度")
        
        # 打印每个房间的详细记录
        print("\n" + "=" * 100)
        print("各房间详细空调记录")
        print("=" * 100)
        
        for room_id in self.room_ids:
            print(f"\n📋 房间 {room_id} 详单:")
            details = self.client.get_ac_details(room_id).get("data", {})
            records = details.get("details", [])
            
            if not records:
                print("  (无记录)")
                continue
            
            print(f"  {'序号':<4} {'开始时间':<20} {'时长(秒)':<10} {'起始温度':<10} {'目标温度':<10} {'风速':<6} {'能耗':<8} {'费用':<8}")
            print("  " + "-" * 90)
            
            for r in records:
                print(f"  {r['seq']:<4} {r['start_time']:<20} {r['duration_seconds']:<10} "
                      f"{r['start_temp']:<10.1f} {r['target_temp']:<10.1f} {r['fan_speed']:<6} "
                      f"{r['energy']:<8.2f} {r['cost']:<8.2f}")
    
    def run_test(self):
        """运行测试"""
        # 解析测试数据
        print(f"📂 加载测试数据: {self.test_data_file}")
        test_data = parse_test_data(self.test_data_file)
        print(f"✅ 共解析 {len(test_data)} 个时间点\n")
        
        # 初始化
        self.setup()
        
        print("=" * 60)
        print("开始执行测试")
        print(f"⏱️  时间压缩比: {TIME_SCALE}x (10秒测试 = 60秒系统时间)")
        print("=" * 60)
        
        self.test_start_time = time.time()
        
        for time_min, actions in test_data:
            # 等待到指定时间点
            target_test_time = time_min * TEST_INTERVAL
            current_test_time = time.time() - self.test_start_time
            
            if target_test_time > current_test_time:
                wait_time = target_test_time - current_test_time
                print(f"\n⏳ 等待 {wait_time:.1f} 秒到达时间点 {time_min} 分钟...")
                time.sleep(wait_time)
            
            print(f"\n{'='*60}")
            print(f"⏰ 时间点: {time_min} 分钟 (已运行: {current_test_time:.1f}秒)")
            print(f"{'='*60}")
            
            if time_min == 0:
                print("  🎬 系统启动，设置制热模式")
                continue
            
            # 执行操作
            if actions:
                print("  📝 执行操作:")
                for room_id, action in actions.items():
                    print(f"    [{room_id}] ", end="")
                    self.execute_action(room_id, action)
            else:
                print("  (无操作)")
            
            # 打印状态
            self.print_status(time_min)
        
        # 测试结束
        self.print_final_report()
        
        print("\n✅ 测试执行完毕！")


# ============================================================
# 主函数
# ============================================================

def main():
    """主函数"""
    print("=" * 60)
    print("制热模式API测试驱动器")
    print(f"启动时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"后端地址: {API_BASE_URL}")
    print("=" * 60)
    
    # 检查后端是否可访问
    try:
        response = requests.get(f"{API_BASE_URL}/rooms/", timeout=5)
        if response.status_code != 200:
            print("❌ 后端API无法访问，请确保Django服务正在运行")
            sys.exit(1)
        print("✅ 后端API连接正常\n")
    except requests.exceptions.RequestException as e:
        print(f"❌ 无法连接后端: {e}")
        sys.exit(1)
    
    # 创建客户端和测试实例
    client = APIClient(API_BASE_URL)
    test = HeatingAPITest(client, TEST_DATA_FILE)
    
    try:
        test.run_test()
    except KeyboardInterrupt:
        print("\n\n⚠️  测试被用户中断")
    except Exception as e:
        print(f"\n\n❌ 测试执行出错: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()